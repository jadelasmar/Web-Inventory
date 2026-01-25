"""Inventory view page."""
from io import BytesIO
from datetime import datetime

import pandas as pd
import streamlit as st
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle

from core.services import (
    get_products,
    get_latest_purchase_parties,
    find_product_by_name,
    add_product,
    update_product,
    record_movement,
    get_product_movement_summary,
    upsert_initial_stock,
)
from ui.components import maybe_open_image_modal, render_products_table


def render(conn):
    """Render the inventory page."""
    st.header("\U0001F5C2\ufe0f Inventory")
    df = get_products(conn)
    party_map = get_latest_purchase_parties(conn)
    if not df.empty:
        df = df.copy()
        df["party"] = df["name"].map(party_map).fillna("")

    search = st.text_input("Search by name, category, brand, description, or party")
    if not df.empty:
        # Case-insensitive partial search across name, category, brand, description, party
        if search:
            mask = df["name"].str.contains(search, case=False, na=False, regex=False)
            if "category" in df.columns:
                mask = mask | df["category"].str.contains(search, case=False, na=False, regex=False)
            if "brand" in df.columns:
                mask = mask | df["brand"].str.contains(search, case=False, na=False, regex=False)
            if "description" in df.columns:
                mask = mask | df["description"].str.contains(search, case=False, na=False, regex=False)
            if "party" in df.columns:
                mask = mask | df["party"].str.contains(search, case=False, na=False, regex=False)
            df = df[mask]
        df = (
            df.sort_values("name", key=lambda s: s.str.casefold())
            if "name" in df.columns
            else df
        )
    maybe_open_image_modal()
    render_products_table(df)

    if st.session_state.get("admin_mode"):
        st.divider()
        st.subheader("Import / Export")

        template_cols = [
            "Name",
            "Category",
            "Brand",
            "Stock",
            "Cost",
            "Price",
            "Description",
            "Image URL",
        ]
        template_df = pd.DataFrame(columns=template_cols + ["__placeholder__"])
        template_df = template_df.drop(columns=["__placeholder__"])
        template_df.loc[0] = [""] * len(template_cols)
        template_buf = BytesIO()
        with pd.ExcelWriter(template_buf, engine="openpyxl") as writer:
            template_df.to_excel(writer, index=False, sheet_name="products")
            ws = writer.sheets["products"]
            max_col = len(template_cols)
            if max_col:
                from openpyxl.utils import get_column_letter
                from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo as XlTableStyleInfo
                last_col = get_column_letter(max_col)
                table = XlTable(displayName="ProductsTemplate", ref=f"A1:{last_col}2")
                style = XlTableStyleInfo(
                    name="TableStyleMedium9",
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False,
                )
                table.tableStyleInfo = style
                ws.add_table(table)
                for idx, col_name in enumerate(template_cols, start=1):
                    ws.column_dimensions[get_column_letter(idx)].width = max(12, len(col_name) + 2)
        st.download_button(
            "Download Excel template",
            data=template_buf.getvalue(),
            file_name="inventory_template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        upload = st.file_uploader(
            "Import from Excel",
            type=["xlsx", "xls"],
            help="Use the template headers: Name, Category, Brand, Stock, Cost, Price, Description, Image URL",
        )
        st.caption("Name, Category, and Brand are required. Use 'Other' if needed.")

        if upload is not None:
            try:
                import_df = pd.read_excel(upload)
            except Exception as e:
                st.error(f"Failed to read file: {e}")
                return

            friendly_map = {
                "name": "product_name",
                "product name": "product_name",
                "category": "category",
                "brand": "brand",
                "stock": "stock",
                "cost": "cost",
                "price": "price",
                "description": "description",
                "image url": "image_url",
                "image_url": "image_url",
                "product_name": "product_name",
            }
            normalized = {}
            for col in import_df.columns:
                key = col.lower().strip()
                normalized[col] = friendly_map.get(key, key)
            import_df = import_df.rename(columns=normalized)

            required_cols = [
                "product_name",
                "category",
                "brand",
                "stock",
                "cost",
                "price",
                "description",
                "image_url",
            ]
            missing = [c for c in required_cols if c not in import_df.columns]
            if missing:
                st.error(f"Missing columns: {', '.join(missing)}")
                return
            import_df = import_df[required_cols]
            import_df = import_df.fillna("")
            st.dataframe(import_df.head(25), use_container_width=True, hide_index=True)

            if st.button("Import products"):
                current_df = get_products(conn, include_inactive=True)
                current_by_name = {}
                if not current_df.empty and "name" in current_df.columns:
                    for _, row in current_df.iterrows():
                        current_by_name[str(row["name"]).strip().lower()] = row

                added, updated, skipped, errors = 0, 0, 0, 0
                error_rows = []
                for idx, row in import_df.iterrows():
                    name = str(row["product_name"]).strip()
                    if not name:
                        skipped += 1
                        error_rows.append(
                            {"row_index": int(idx) + 2, "product_name": "", "error": "Missing Name"}
                        )
                        continue

                    key = name.lower()
                    category = str(row["category"]).strip()
                    brand = str(row["brand"]).strip()
                    description = str(row["description"]).strip()
                    image_url = str(row["image_url"]).strip()
                    if not image_url:
                        from pathlib import Path
                        image_dir = Path("assets/product_images")
                        for ext in [".png", ".jpg", ".jpeg", ".gif", ".webp"]:
                            candidate = image_dir / f"{name}{ext}"
                            if candidate.exists():
                                image_url = str(candidate)
                                break
                    cost_raw = str(row["cost"]).strip()
                    price_raw = str(row["price"]).strip()
                    stock_raw = str(row["stock"]).strip()

                    cost = None if cost_raw == "" else None
                    price = None if price_raw == "" else None
                    stock = None if stock_raw == "" else None

                    if cost_raw != "":
                        try:
                            cost = float(cost_raw)
                        except Exception:
                            errors += 1
                            error_rows.append({
                                "row_index": int(idx) + 2,
                                "product_name": name,
                                "error": f"Invalid cost: '{cost_raw}'",
                            })
                            continue
                    else:
                        cost = 0.0

                    if price_raw != "":
                        try:
                            price = float(price_raw)
                        except Exception:
                            errors += 1
                            error_rows.append({
                                "row_index": int(idx) + 2,
                                "product_name": name,
                                "error": f"Invalid price: '{price_raw}'",
                            })
                            continue
                    else:
                        price = 0.0

                    if stock_raw != "":
                        try:
                            stock = int(float(stock_raw))
                        except Exception:
                            errors += 1
                            error_rows.append({
                                "row_index": int(idx) + 2,
                                "product_name": name,
                                "error": f"Invalid stock: '{stock_raw}'",
                            })
                            continue

                    try:
                        if not category or not brand:
                            errors += 1
                            missing_parts = []
                            if not category:
                                missing_parts.append('Category')
                            if not brand:
                                missing_parts.append('Brand')
                            missing_label = ' and '.join(missing_parts)
                            error_rows.append(
                                {
                                    "row_index": int(idx) + 2,
                                    "product_name": name,
                                    "error": f"Missing {missing_label} (use 'Other')",
                                }
                            )
                            continue
                        existing = current_by_name.get(key)
                        if existing is None:
                            add_product(
                                conn,
                                (
                                    name,
                                    category,
                                    brand,
                                    description,
                                    image_url,
                                    0,
                                    cost,
                                    price,
                                    "",
                                ),
                            )
                            added += 1
                        else:
                            update_product(
                                conn,
                                (
                                    name,
                                    category,
                                    brand,
                                    description,
                                    image_url,
                                    cost,
                                    price,
                                    existing.get("supplier", ""),
                                    existing.get("name"),
                                ),
                            )
                            updated += 1

                        if stock is not None:
                            summary = get_product_movement_summary(conn, name)
                            total_movements = summary.get("total_count", 0)
                            has_only_initial = (
                                total_movements == 1
                                and summary.get("initial_stock_id") is not None
                            )
                            if total_movements == 0 or has_only_initial:
                                upsert_initial_stock(
                                    conn,
                                    product_name=name,
                                    quantity=int(stock),
                                    price=cost,
                                    supplier_customer="",
                                    notes="Imported initial stock",
                                    movement_date=datetime.now().date(),
                                    movement_id=summary.get("initial_stock_id")
                                    if has_only_initial
                                    else None,
                                )
                            else:
                                current_stock = int(existing.get("current_stock", 0)) if existing is not None else 0
                                delta = int(stock) - current_stock
                                if delta != 0:
                                    record_movement(
                                        conn,
                                        (
                                            name,
                                            category,
                                            "ADJUSTMENT",
                                            int(delta),
                                            "N/A",
                                            "",
                                            "Imported stock adjustment",
                                            datetime.now().date(),
                                        ),
                                    )
                    except Exception as e:
                        errors += 1
                        error_rows.append(
                            {
                                "row_index": int(idx) + 2,
                                "product_name": name,
                                "error": f"Unexpected error: {e}",
                            }
                        )

                st.success(
                    f"Import complete. Added: {added}, Updated: {updated}, Skipped: {skipped}, Errors: {errors}"
                )
                if error_rows:
                    error_df = pd.DataFrame(error_rows)
                    error_buf = BytesIO()
                    with pd.ExcelWriter(error_buf, engine="openpyxl") as writer:
                        error_df.to_excel(writer, index=False, sheet_name="import_errors")
                    st.download_button(
                        "Download import error report",
                        data=error_buf.getvalue(),
                        file_name="inventory_import_errors.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

        # Export
        if not df.empty:
            export_cols = [
                "name",
                "category",
                "brand",
                "current_stock",
                "cost_price",
                "sale_price",
                "party",
                "description",
                "image_url",
            ]
            export_df = df.copy()
            for col in export_cols:
                if col not in export_df.columns:
                    export_df[col] = ""
            export_df = export_df[export_cols].rename(
                columns={
                    "name": "Name",
                    "category": "Category",
                    "brand": "Brand",
                    "current_stock": "Stock",
                    "cost_price": "Cost",
                    "sale_price": "Price",
                    "party": "Supplier",
                    "description": "Description",
                    "image_url": "Image URL",
                }
            )

            excel_buf = BytesIO()
            with pd.ExcelWriter(excel_buf, engine="openpyxl") as writer:
                export_df.to_excel(writer, index=False, sheet_name="products")
                ws = writer.sheets["products"]
                max_col = len(export_df.columns)
                max_row = len(export_df) + 1
                if max_col:
                    from openpyxl.utils import get_column_letter
                    from openpyxl.worksheet.table import Table as XlTable, TableStyleInfo as XlTableStyleInfo
                    last_col = get_column_letter(max_col)
                    table = XlTable(
                        displayName="ProductsExport",
                        ref=f"A1:{last_col}{max_row}"
                    )
                    style = XlTableStyleInfo(
                        name="TableStyleMedium9",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False,
                    )
                    table.tableStyleInfo = style
                    ws.add_table(table)
                    for idx, col_name in enumerate(export_df.columns, start=1):
                        ws.column_dimensions[get_column_letter(idx)].width = max(12, len(str(col_name)) + 2)
            st.download_button(
                "Export to Excel",
                data=excel_buf.getvalue(),
                file_name="inventory_export.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

            pdf_buf = BytesIO()
            doc = SimpleDocTemplate(pdf_buf, pagesize=landscape(letter))
            pdf_headers = list(export_df.columns)
            data = [pdf_headers] + export_df.fillna("").astype(str).values.tolist()
            table = Table(data)
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ]
                )
            )
            doc.build([table])
            st.download_button(
                "Export to PDF",
                data=pdf_buf.getvalue(),
                file_name="inventory_export.pdf",
                mime="application/pdf",
            )
